#!/bin/sh

''':'
exec $(dirname $0)/../bin/python "$0" "$@"
'''

def get_cell_ref(ws, start_col=1, end_col=None, start_row=1, end_row=None):
    from openpyxl.utils.cell import get_column_letter
    end_col = end_col or ws.max_column
    end_row = end_row or ws.max_row
    return "{}{}:{}{}".format(get_column_letter(start_col), start_row, get_column_letter(end_col), end_row)

def set_table_style(ws, name="Sheet", start_col=1, end_col=None, start_row=1, end_row=None):
    from openpyxl.worksheet.table import Table, TableStyleInfo
    name = name.replace(" ", "").strip()
    end_col = end_col or ws.max_column
    end_row = end_row or ws.max_row
    ref = get_cell_ref(ws, start_col, end_col, start_row, end_row)
    tab = Table(displayName="Table{}".format(name), ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def create_workbook():
    from openpyxl import Workbook
    wb = Workbook()
    del wb['Sheet']
    return wb

def create_sheet(wb, name, cols, rows, addl_cols=None):

    ws = wb.create_sheet(name)

    # start adding rows and columns
    start_row, start_col = 1, 1

    # add addtional header
    if addl_cols:
        row, merges = [], []
        for col, span in addl_cols:
            row.append(col)
            span = int(span)-1
            if span > 0:
                start_col0 = start_col + len(row) - 1
                col_index = len(row) - 1
                row.extend([""]*span)
                end_col = start_col + len(row) - 1
                ref = get_cell_ref(ws, start_col0, end_col, col_index, col_index)
                merges.append(ref)
        ws.append(row)
        for merge in merges:
            ws.merge_cells(merge)
        start_row = start_row + 1

    # add columns
    ws.append(cols)

    # add rows
    for row in rows:
        ws.append(row)

    # set style
    set_table_style(ws, name)

    return ws


def write_xls_file(cols, rows, filepath=None, addl_cols=None):
    wb = create_workbook()
    create_sheet(wb, "Sheet1", cols, rows, addl_cols=addl_cols)
    wb.save(filepath or "default.xlsx")

def read_xls_file(filepath):
    import openpyxl
    wb = openpyxl.load_workbook(filepath)
    #print (wb.sheetnames)
    rv = []
    for ws in wb.worksheets:
        cols = []
        for i in range(1, ws.max_column + 1):
            cols.append(ws.cell(row = 1, column = i).value)

        rows = []
        for i in range(1, ws.max_row + 1):
            row = []
            for cell in ws[i]:
                row.append(cell.value)
            rows.append(row)
        rv.append([cols, rows])
    return rv

def unit_test():
    cols = ["C1", "C2", "C3"]
    rows = [
             ["r11", "r12", "r13"],
             ["r21", "r22", "r23"],
             ["r31", "r32", "r33"]
           ]
    addl_cols = [["H0", "1"], ["H1", "2"]]
    write_xls_file(cols, rows, "unit-test.xlsx", addl_cols=addl_cols)
    wb = create_workbook()
    create_sheet(wb, "Sheet1", cols, rows, addl_cols=addl_cols)
    create_sheet(wb, "Sheet2", cols, rows)
    wb.save("unit-test-3.xlsx")
    print(read_xls_file("unit-test-3.xlsx")[0])

if __name__ == "__main__":
    unit_test()
