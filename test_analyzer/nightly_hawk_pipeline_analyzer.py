#!/bin/env python
'''testbed auto recovery
1: collect unhealthy testbeds
2: power cycle all testbeds
3: ping testbeds one by one again
4: redeploy testbed
5: sanity check
6: upload unhealthy table to Custo
'''

from __future__ import print_function, division

import argparse
import json
import logging
import os
import sys
import requests
import time
import datetime
import yaml
import re

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter 
from openpyxl.styles import Font, colors, Alignment, PatternFill, Border, Side

# from dateutil.tz import tzutc

from nightly_hawk_autoRecovery_cfg import curr_convert_to_trusty_images_dict
from nightly_hawk_autoRecovery_cfg import nightly_pipeline_internal
from nightly_hawk_autoRecovery_cfg import nightly_pipeline_master
from nightly_hawk_autoRecovery_cfg import nightly_pipeline_202012
from nightly_hawk_autoRecovery_cfg import nightly_pipeline_202205

from nightly_hawk_common import NightlyPipelineCheck, TbShare, KustoChecker, KustoUploader

NIGHTLY_HAWK_DIR = os.path.abspath(os.path.dirname(__file__))
SONIC_MGMT_DIR = os.path.dirname(NIGHTLY_HAWK_DIR)
ANSIBLE_DIR = os.path.join(SONIC_MGMT_DIR, 'ansible') 
NIGHTLY_PIPELINE_YML_DIR = os.path.join(SONIC_MGMT_DIR, '.azure-pipelines/nightly') 


logging.basicConfig(
    stream=sys.stdout,
    level=logging.INFO,
    format='%(asctime)s %(filename)s:%(name)s:%(lineno)d %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class Nightly_hawk_pipeline_analyzer(object):
    def __init__(self, verbose = False, days2parser = 3):
        self.verbose = verbose

        self.pipeline_parser_analyzer_dict = {}
        self.nightly_pipelines_yml_dict = {}
        self.pipeline_trusty_images = {}
        self.nightly_pipelines_dict = {}
        self.nightly_pipeline_testbeds = []

        self.pipeline_build_result_dict = {}
        self.build_test_case_result_dict = {}
        
        self.pipeline_cols = 2
        self.pipeline_rows = 14
        self.pipeline_cols_name_width = 18
        self.pipeline_cols_data_width = 70
        self.pipeline_date_list = []
        # for i in range(self.pipeline_date_parser_count):
        for i in range(days2parser):
            self.pipeline_date_list.append(str(datetime.date.today() - datetime.timedelta(days=i)))
        self.save_file_name = str(datetime.date.today()) + "_pipeline_parser.xlsx"

        self.DATABASE = 'SonicTestData'
        self.kusto_checker = self.create_kusto_checker()
        self.nightly_pipeline_check = NightlyPipelineCheck()
        self.pipeline_build_case_result = {}

        self.result_value_list = ['success', 'error', 'failure', 'skipped', 'xfail_expected', 'xfail_unexpected', 'xfail_forgive']
        self.keys = ["StartTimeUTC", "TestbedName", "OSVersion", "Result", "BuildId", "FullTestPath", "Comments", "Summary", "StartTime", "Runtime", "ModulePath", "TestCase"]
        

    def create_kusto_checker(self):
        # ingest_cluster = os.getenv("TEST_REPORT_INGEST_KUSTO_CLUSTER")
        # cluster = ingest_cluster.replace('ingest-', '')
        # tenant_id = os.getenv("TEST_REPORT_AAD_TENANT_ID")
        # client_id = os.getenv("TEST_REPORT_AAD_CLIENT_ID")
        # client_key = os.getenv("TEST_REPORT_AAD_CLIENT_KEY")

        ingest_cluster = os.getenv("TEST_REPORT_INGEST_KUSTO_CLUSTER_BACKUP")
        cluster = ingest_cluster.replace('ingest-', '')
        tenant_id = os.getenv("TEST_REPORT_AAD_TENANT_ID_BACKUP")
        client_id = os.getenv("TEST_REPORT_AAD_CLIENT_ID_BACKUP")
        client_key = os.getenv("TEST_REPORT_AAD_CLIENT_KEY_BACKUP")

        if not all([cluster, tenant_id, client_id, client_key]):
            raise RuntimeError('Could not load Kusto credentials from environment')

        return KustoChecker(cluster, tenant_id, client_id, client_key, self.DATABASE)


    def sorted_key_dict(self, myDict):
        myKeys = list(myDict.keys())
        myKeys.sort()
        sorted_dict = {i: myDict[i] for i in myKeys}
        return sorted_dict

    def sorted_value_dict(self, myDict):
        myList=sorted((value, key) for (key,value) in myDict.items())
        logger.info("myList {} ".format(myList))

        sorted_dict=dict([(k,v) for v,k in myList])
        logger.info("sorted_dict {} ".format(sorted_dict))
        return sorted_dict

    def collect_nightly_pipeline_build_by_branch(self, branch):
        if 'master' == branch :
            nightly_pipeline_branch_dict = nightly_pipeline_master.copy()
        elif 'internal' == branch :
            nightly_pipeline_branch_dict = nightly_pipeline_internal.copy()
        elif 'internal-202012' == branch :
            nightly_pipeline_branch_dict = nightly_pipeline_202012.copy()
        elif 'internal-202205' == branch :
            nightly_pipeline_branch_dict = nightly_pipeline_202205.copy()
        else:
            logger.error("ERROR: branch {} mismatch ".format(branch))
            raise RuntimeError('collect_pipeline_build_images, branch {} mismatch'.format(branch))

        if branch not in self.nightly_pipelines_dict.keys():
            self.nightly_pipelines_dict[branch] = {}

        for testbed, pipeline_info in self.pipeline_parser_analyzer_dict.items(): 
            branch_info = pipeline_info.get(branch, None)
            # logger.info("testbed {} pipeline_info {}  ".format(testbed, pipeline_info)) 
            if not branch_info:
                continue
            for yml_name, pipeline_info in branch_info.items(): 
                # logger.info("yml_name {} pipeline_info {}  ".format(yml_name, pipeline_info))
                for index, pipeline_name in nightly_pipeline_branch_dict.items(): 
                    # logger.info("index {} pipeline_name {}  ".format(index, pipeline_name))
                    if pipeline_info['pipeline_name'] == pipeline_name:
                        # logger.info("testbed {} branch_info {}  ".format(testbed, branch_info))
                        self.nightly_pipelines_dict[branch][index] = {
                                                                        'pipeline_name' : pipeline_info['pipeline_name'],
                                                                        'testbed_name' : testbed,
                                                                        'image_url' : pipeline_info['image_url'],
                                                                        'schedule' : pipeline_info['schedule'],
                                                                        'pipeline_id' : pipeline_info['pipeline_id'],
                                                                     }

                        if testbed not in self.nightly_pipeline_testbeds:
                            self.nightly_pipeline_testbeds.append(testbed)

                        break
        
        logger.info(" nightly_pipelines_dict {}  ".format(self.nightly_pipelines_dict))




    def collect_nightly_pipeline_build(self):
        """
        followed nightly build result table in PowerBI, nightly_pipeline_202012

        self.nightly_pipelines_dict
        'internal': {
            1: {
            'pipeline_name': 'vms1-t1-2700-platform_tests.internal',
            'testbed_name': 'vms1-t1-2700',
            'image_url': '$(IMAGE_MLNX_INTERNAL)',
            'schedule': '10 4 * * 1,3',
            'pipeline_id': 505
            },
            2: {
            'pipeline_name': 'vms11-2-t0-2700-2-platform_tests.internal',
            'testbed_name': 'vms11-2-t0',
            'image_url': '$(IMAGE_MLNX_INTERNAL)',
            'schedule': '0 4 * * 1,3',
            'pipeline_id': 520
            },  
            ......   
        """
        self.collect_nightly_pipeline_build_by_branch('master')
        # logger.info("nightly master {}  ".format(self.sorted_key_dict(self.nightly_pipelines_dict['master'])))

        self.collect_nightly_pipeline_build_by_branch('internal')
        # logger.info("nightly internal {}  ".format(self.sorted_key_dict(self.nightly_pipelines_dict['internal'])))

        self.collect_nightly_pipeline_build_by_branch('internal-202012')
        # logger.info("nightly 202012 {}  ".format(self.sorted_key_dict(self.nightly_pipelines_dict['internal-202012'])))

        self.collect_nightly_pipeline_build_by_branch('internal-202205')
        # logger.info("nightly 202205 {}  ".format(self.sorted_key_dict(self.nightly_pipelines_dict['internal-202205'])))

        # logger.info("nightly {}  ".format(self.sorted_key_dict(self.nightly_pipelines_dict)))
        # logger.info("nightly_pipeline_testbeds {}  ".format(self.nightly_pipeline_testbeds))

        # self.collect_pipeline_trusty_images()
        # logger.info("pipeline_trusty_images {}  ".format(self.pipeline_trusty_images))

    def collect_pipeline_trusty_images(self):
        """
        poll and collect all pipelines' image, and convert/save it 
        """
        for testbed, pipeline_build in self.pipeline_parser_analyzer_dict.items(): 
            logger.info("testbed {} pipeline_build {}  ".format(testbed, pipeline_build))
            self.pipeline_trusty_images[testbed] = None
            if pipeline_build.get('internal-202012', None) :
                for pipeline_yml, pipeline_info in pipeline_build['internal-202012'].items(): 
                    if pipeline_info['image_url'] and not self.pipeline_trusty_images[testbed]:
                        logger.info("202012: testbed {} pipeline_build {}  ".format(testbed, pipeline_info['image_url']))
                        self.pipeline_trusty_images[testbed] = pipeline_info['image_url']
            elif pipeline_build.get('internal-202205', None) :
                for pipeline_yml, pipeline_info in pipeline_build['internal-202205'].items(): 
                    if pipeline_info['image_url'] and not self.pipeline_trusty_images[testbed]:
                        logger.info("202205: testbed {} pipeline_build {}  ".format(testbed, pipeline_info['image_url']))
                        self.pipeline_trusty_images[testbed] = pipeline_info['image_url']
            elif pipeline_build.get('internal', None) :
                for pipeline_yml, pipeline_info in pipeline_build['internal'].items(): 
                    if pipeline_info['image_url'] and not self.pipeline_trusty_images[testbed]:
                        logger.info("internal: testbed {} pipeline_build {}  ".format(testbed, pipeline_info['image_url']))
                        self.pipeline_trusty_images[testbed] = pipeline_info['image_url']
            elif pipeline_build.get('master', None) :
                for pipeline_yml, pipeline_info in pipeline_build['master'].items(): 
                    if pipeline_info['image_url'] and not self.pipeline_trusty_images[testbed]:
                        logger.info("master: testbed {} pipeline_build {}  ".format(testbed, pipeline_info['image_url']))
                        self.pipeline_trusty_images[testbed] = pipeline_info['image_url']

            if not self.pipeline_trusty_images[testbed]:
                logger.warning("testbed {} has no image url ".format(testbed))
                del self.pipeline_trusty_images[testbed]

        logger.info("pipeline_trusty_images {}  ".format(self.pipeline_trusty_images))

        curr_images_list = []
        for curr_image in self.pipeline_trusty_images.values():
            if curr_image and curr_image not in curr_images_list:
                curr_images_list.append(curr_image)
        logger.info("pipeline current images list {}  ".format(curr_images_list))

        for testbed, trusty_image in self.pipeline_trusty_images.items(): 
            if trusty_image:
                if curr_convert_to_trusty_images_dict.get(trusty_image, None):
                    self.pipeline_trusty_images[testbed] = curr_convert_to_trusty_images_dict[trusty_image]
                else:
                    logger.error("testbed {}  image {} has no trusty image ".format(testbed, trusty_image))

        logger.info("pipeline trusty images {}  ".format(self.pipeline_trusty_images))
        # logger.info("pipeline sorted trusty images {}  ".format(self.sorted_value_dict(self.pipeline_trusty_images)))


    def get_pipelines_build_log_buffer(self, url):
        TOKEN = os.environ.get('AZURE_DEVOPS_MSSONIC_TOKEN')
        if not TOKEN:
            logger.error("Get token failed, Must export environment variable AZURE_DEVOPS_MSSONIC_TOKEN")
            if self.verbose :
                logger.error("No token ")
        AUTH = ('', TOKEN)

        for i in range(3):
            build_response = requests.get(url, auth=AUTH)
            if build_response.status_code == requests.codes.ok:
                break
            time.sleep(1)

        if build_response.status_code == requests.codes.ok:
            return build_response.text
        else:
            logger.error("get_pipelines_build_log_buffer failed: code {}, url {} ".format(build_response.status_code, url))
            return None

    def parser_run_tests_logs(self, log_buffer):
        failure_output = ''

        tmp_log_file = "run_test.log"
        with open(tmp_log_file, "w", encoding='utf-8') as out_file: 
            out_file.write(log_buffer)

        last_run_case = {'pass' : {'pattern' : '.*PASSED\s+\[.*\d\%\]', 'case' : None}, 
                         'skip' : {'pattern' : '.*SKIPPED\s+\[.*\d\%\]', 'case' : None}, 
                         'error' : {'pattern' : '.*ERROR\s+\[.*\d\%\]', 'case' : None}}
        f = open("run_test.log","r")
        lines = f.readlines()
        line_number = 0
        pre_line_data = None
        for line in lines:
            line_number += 1
            # logger.info("line_number {} line {}   ".format(line_number, line))
            if 'PASSED' in line:
                last_pass_case = re.findall(last_run_case['pass']['pattern'], line)
                if len(last_pass_case) > 0 :
                    if 'test_collect_techsupport' in last_pass_case[0] \
                        or 'test_restore_container_autorestart' in last_pass_case[0] \
                        or 'test_recover_rsyslog_rate_limit' in last_pass_case[0]:

                        continue
                    if len(re.findall('\d+-\d+-\d+T\d+:\d+:\d+\.\d+Z PASSED', last_pass_case[0])) > 0:
                        last_run_case['pass']['case'] = pre_line_data + last_pass_case[0]
                    else:
                        last_run_case['pass']['case'] = last_pass_case[0]
                    # logger.info("line_number {} pass {}   ".format(line_number, last_run_case['pass']['case'] ))
            elif 'ERROR' in line:
                last_error_case = re.findall(last_run_case['error']['pattern'], line)
                if len(last_error_case) > 0 :
                    if len(re.findall('\d+-\d+-\d+T\d+:\d+:\d+\.\d+Z ERROR', last_error_case[0])) > 0:
                        last_run_case['error']['case'] = pre_line_data + last_error_case[0]
                    else:
                        last_run_case['error']['case'] = last_error_case[0]
                    # logger.info("line_number {} error {}   ".format(line_number, last_run_case['error']['case'] ))
            elif 'SKIPPED' in line:
                last_skip_case = re.findall(last_run_case['skip']['pattern'], line)
                if len(last_skip_case) > 0 :
                    if len(re.findall('\d+-\d+-\d+T\d+:\d+:\d+\.\d+Z SKIPPED', last_skip_case[0])) > 0:
                        last_run_case['skip']['case'] = pre_line_data + last_skip_case[0]
                    else:
                        last_run_case['skip']['case'] = last_skip_case[0]
                    # logger.info("line_number {} skip {}   ".format(line_number, last_run_case['skip']['case'] ))

            pre_line_data = line

        failure_output = '{}\nlast pass cases; {}'.format(failure_output, last_run_case['pass']['case'])
        # failure_output = '{}\nlast skip cases; {}'.format(failure_output, last_run_case['skip']['case'])
        failure_output = '{}\nlast error cases; {}'.format(failure_output, last_run_case['error']['case'])

        return failure_output


    def pipelines_build_log_match(self, buffer, pattern_list):
        """
        search the buffer and return whether has input pattern
        """
        matched_patter = ""
        for i in range(len(pattern_list)):
            find_result = re.findall(pattern_list[i], buffer)
            if len(find_result) > 0:
                matched_patter = '{}\n{}'.format(matched_patter, find_result[0])
        return matched_patter

    def pipelines_build_log_serach(self, buffer, pattern):
        """
        search the pattern in buffer and return the whole line which includs the pattern
        """
        searched_pattern = ""
        with open("tmp.txt", "w", encoding='utf-8') as out_file: 
            out_file.write(buffer)
        f = open("tmp.txt","r")
        lines = f.readlines()
        line_number = 0
        for line in lines:
            line_number += 1
            if pattern in line:
                skip = re.compile(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')
                searched_pattern = skip.sub('', line)
                break
        return searched_pattern

    def collect_pipeline_build_case_result_by_build_id(self, build_id):
        """
        Collects test results from Kusto based on build ID.
        :param build_id: str - the ID of the build to retrieve test results for.
        """
        build_test_result = self.kusto_checker.query_build_test_result(build_id)
        logger.debug("build_test_result {} ".format(build_test_result)) 

        # result_header_list = ['StartTimeUTC', 'TestbedName', 'OSVersion', 'StartTime', 'Runtime', 'Result', 'BuildId', 'ModulePath', 'TestCase', 'FullTestPath', 'Summary']
        # result_value_list = ['success', 'error', 'failure', 'skipped', 'xfail_expected', 'xfail_forgive']
        result_dict = {result: {} for result in self.result_value_list}

        # keys = ["StartTimeUTC", "TestbedName", "OSVersion", "Result", "BuildId", "FullTestPath", "Comments" "Summary", "StartTime", "Runtime", "ModulePath", "TestCase"]
        index = 0
        for row in build_test_result.primary_results[0].rows:
            # logger.debug("index {} ".format(index)) 
            if row['Result'] not in self.result_value_list:
                logger.error("build id {} case {} result {} not expected".format(build_id, row['FullTestPath'], row['Result'])) 
            else:
                result_dict[row['Result']][index] = {}
                # logger.info("row {} ".format(row))
                for i, key in enumerate(self.keys):
                    # logger.info("{}; key: {}, value: {}; ".format(i, key, str((list(row))[i])))
                    result_dict[row['Result']][index][key] = str((list(row))[i])
            index += 1

        logger.info("Retrieved {} test results for build ID {} ".format(len(build_test_result.primary_results[0].rows), build_id))

        for value in self.result_value_list:
            logger.debug("{} cases count: {} ".format(value, len(result_dict[value].keys())))

        return result_dict

    def create_branch_verify_result_xls_file(self):
        logger.info("save_file_name {}".format(self.save_file_name)) 
        if os.path.exists(self.save_file_name):
            os.system("rm -f {}".format(self.save_file_name))

        wb = Workbook()

        for build_id in self.branch_verify_result.keys():
            logger.debug("build_id {} ".format(build_id)) 
            self.build_id_result_excel_sheet_create(wb, build_id, self.branch_verify_result[build_id] )

        wb.save(self.save_file_name)
        return
    

    def parser_pipeline_build_result_by_build_id(self, pipeline_id, build_id):
        get_build_records = self.nightly_pipeline_check.get_pipeline_build_result_by_build_id(pipeline_id, build_id)
        if get_build_records == None:
            return

        logger.debug("get_pipeline_build_result_by_build_id {}  ".format((get_build_records)))

        parser_items = ['Lock Testbed', 'Upgrade Image', 'Deploy Minigraph', 'Run Tests']
        for i in range(int(len(get_build_records['records']))):
            for j in range(len(parser_items)) :
                if parser_items[j] in get_build_records['records'][i].values() :
                    if get_build_records['records'][i]['state'] == 'completed':
                        logger.debug("pipeline_id {} build_id {} record {}, item {}, result {}".format(pipeline_id, build_id, i, parser_items[j], get_build_records['records'][i]['result'])) 

                        self.pipeline_build_result_dict[pipeline_id][build_id]['build_details'][parser_items[j]] = get_build_records['records'][i]['result']
                        if get_build_records['records'][i]['result'] == 'failed' :
                            if parser_items[j] == 'Lock Testbed':
                                pattern = [r'is absolutely locked by (.+?), force lock failed']
                                build_log_buffer = self.get_pipelines_build_log_buffer(get_build_records['records'][i]['log']['url'])
                                testbed_user = self.pipelines_build_log_match(build_log_buffer, pattern)
                                self.pipeline_build_result_dict[pipeline_id][build_id]['build_details'][parser_items[j]] = 'failed\n' + testbed_user
                            elif parser_items[j] == 'Upgrade Image':
                                pattern = 'fatal'
                                build_log_buffer = self.get_pipelines_build_log_buffer(get_build_records['records'][i]['log']['url'])
                                result_output = self.pipelines_build_log_serach(build_log_buffer, pattern)
                                self.pipeline_build_result_dict[pipeline_id][build_id]['build_details'][parser_items[j]] = 'failed\n' + result_output
                            elif parser_items[j] == 'Deploy Minigraph':
                                pattern = 'fatal'
                                build_log_buffer = self.get_pipelines_build_log_buffer(get_build_records['records'][i]['log']['url'])
                                result_output = self.pipelines_build_log_serach(build_log_buffer, pattern)
                                self.pipeline_build_result_dict[pipeline_id][build_id]['build_details'][parser_items[j]] = 'failed\n' + result_output
                            elif parser_items[j] == 'Run Tests':
                                start_time = datetime.datetime.strptime(get_build_records['records'][i]['startTime'][0:19], '%Y-%m-%dT%H:%M:%S')
                                finish_time = datetime.datetime.strptime(get_build_records['records'][i]['finishTime'][0:19], '%Y-%m-%dT%H:%M:%S')
                                run_time = finish_time - start_time
                                logger.debug("record {} item {} run time {} start time {} finish time {} ".format(i, parser_items[j], run_time, get_build_records['records'][i]['startTime'][0:19], get_build_records['records'][i]['finishTime'][0:19])) 

                                pattern_list = ['pretest failed. Please check the detailed log', 'Sanity check failed. Please check the detailed log', 'The task has timed out']
                                build_log_buffer = self.get_pipelines_build_log_buffer(get_build_records['records'][i]['log']['url'])
                                result_output = self.pipelines_build_log_match(build_log_buffer, pattern_list)
                                # result_output = '{}\n{}'.format(result_output, self.parser_run_tests_logs(build_log_buffer))                                
                                self.pipeline_build_result_dict[pipeline_id][build_id]['build_details'][parser_items[j]] = '{}{}\n{}'.format('failed run time: ', run_time, result_output)
                        elif get_build_records['records'][i]['result'] == 'succeeded' :
                            if parser_items[j] == 'Upgrade Image':
                                pattern = 'current image:'
                                build_log_buffer = self.get_pipelines_build_log_buffer(get_build_records['records'][i]['log']['url'])
                                result_output = self.pipelines_build_log_serach(build_log_buffer, pattern)

                                start_index = result_output.find("current image: ")
                                if start_index != -1:
                                    current_image = result_output[start_index:]
                                else:
                                    current_image = "No image found"
                                self.pipeline_build_result_dict[pipeline_id][build_id]['build_details'][parser_items[j]] = current_image

                            elif parser_items[j] == 'Run Tests':
                                build_result_dict = self.collect_pipeline_build_case_result_by_build_id(build_id)
        
                                for value in self.result_value_list:
                                    logger.info("{} cases count: {} ".format(value, len(build_result_dict[value].keys())))

                                self.result_value_list = ['success', 'error', 'failure', 'skipped', 'xfail_expected', 'xfail_unexpected', 'xfail_forgive']

                                count_success_case = len(build_result_dict['success'].keys())
                                count_error_case = len(build_result_dict['error'].keys())
                                count_failure_case = len(build_result_dict['failure'].keys())
                                count_skipped_case = len(build_result_dict['skipped'].keys())
                                count_xfail_expected_case = len(build_result_dict['xfail_expected'].keys())
                                count_xfail_unexpected_case = len(build_result_dict['xfail_unexpected'].keys())
                                count_xfail_forgive_case = len(build_result_dict['xfail_forgive'].keys())
                                count_total_cases = count_success_case + count_error_case + count_failure_case
                                
                                if count_total_cases == 0:
                                    success_rate = 0
                                else:
                                    success_rate = count_success_case / count_total_cases
                                    self.pipeline_build_case_result[build_id] = build_result_dict
                                    
                                # logger.info("success rate {:.2%} ".format(success_rate))
                                start_time = datetime.datetime.strptime(get_build_records['records'][i]['startTime'][0:19], '%Y-%m-%dT%H:%M:%S')
                                finish_time = datetime.datetime.strptime(get_build_records['records'][i]['finishTime'][0:19], '%Y-%m-%dT%H:%M:%S')
                                run_time = finish_time - start_time
                                logger.debug("record {} item {} run time {} start time {} finish time {} ".format(i, parser_items[j], run_time, get_build_records['records'][i]['startTime'][0:19], get_build_records['records'][i]['finishTime'][0:19])) 

                                result = 'success: {}\t error: {}\t failure: {}\t skipped: {}\t xfail_expected: {}\t xfail_unexpected: {}\t xfail_forgive: {}\t pass rate: {:.2%}\t  run time: {} '.format(
                                    count_success_case, count_error_case, count_failure_case, count_skipped_case, count_xfail_expected_case, count_xfail_unexpected_case, count_xfail_forgive_case, success_rate, run_time)
                                logger.info("build_id {} \t result {}".format(build_id, result))
                                logger.info("{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{:.2%}\t{}\t".format(build_id, count_success_case, count_error_case, count_failure_case, count_skipped_case, count_xfail_expected_case, count_xfail_unexpected_case, count_xfail_forgive_case, success_rate, run_time))

                                result_output = 'pass rate: {:.2%}\nsuccess: {}\nerror: {}\nfailure: {}\nskipped: {}\nxfail_expected: {}\nxfail_unexpected: {}\nxfail_forgive: {}\n '.format(
                                    success_rate, count_success_case, count_error_case, count_failure_case, count_skipped_case, count_xfail_expected_case, count_xfail_unexpected_case, count_xfail_forgive_case)
                                # logger.info("build_id {}; success rate {:.2%};  {} ".format(build_id, success_rate, result_output))
                                self.pipeline_build_result_dict[pipeline_id][build_id]['build_details'][parser_items[j]] = '{}{}\n{}'.format('run time: ', run_time, result_output)
        return


    def parser_pipeline_build_result(self):
        for pipeline_id in self.pipeline_build_result_dict.keys():
            logger.info("pipeline_id {}  ".format(pipeline_id)) 
            for build_id in self.pipeline_build_result_dict[pipeline_id].keys():
                logger.info("parser build result pipeline_id {} build_id {} ".format(pipeline_id, build_id))
                for key, value in self.pipeline_build_result_dict[pipeline_id][build_id].items():
                    logger.info("pipeline_id {} build_id {} key {}: {} ".format(pipeline_id, build_id, key, value)) 
                    if key == 'state':
                        state_result = value
                    elif key == 'result':
                        result_value = value
                    # elif key == 'id':
                    #     build_id = value

                logger.info("pipeline_id {} build_id {} state {} result {}  ".format(pipeline_id, build_id, state_result, result_value))
                if state_result == 'completed' :
                    self.parser_pipeline_build_result_by_build_id(pipeline_id, build_id)

        logger.info("pipeline_build_result_dict {}".format(self.pipeline_build_result_dict))

        json_object = json.dumps(self.pipeline_build_result_dict, indent = 4)
        with open("pipeline_build_result_dict_debug.json", "w") as out_file:
            out_file.write(json_object)


    def parser_pipeline_status_result_by_pipeline_id(self, pipeline_id):
        TOKEN = os.environ.get('AZURE_DEVOPS_MSSONIC_TOKEN')
        if not TOKEN:
            logger.error("Get token failed, Must export environment variable AZURE_DEVOPS_MSSONIC_TOKEN")
            if self.verbose :
                logger.error("No token, pipeline_id {} ".format(pipeline_id))
            return
        AUTH = ('', TOKEN)

        pipeline_url = "https://dev.azure.com/mssonic/internal/_apis/pipelines/" + str(pipeline_id) + "/runs?api-version=6.0-preview.1"

        for i in range(3):
            build_response = requests.get(pipeline_url, auth=AUTH)
            if build_response.status_code == requests.codes.ok:
                break
            time.sleep(1)

        if build_response.status_code == requests.codes.ok:
            pipeline_dict = yaml.safe_load(build_response.text)
            # logger.info("pipeline_id {} dict {}  ".format(pipeline_id, pipeline_dict))


            parser_count = pipeline_dict['count']
            logger.debug("pipeline_id {} parser_count {} ".format(pipeline_id, parser_count))

            if not self.pipeline_build_result_dict.get(pipeline_id, None):
                self.pipeline_build_result_dict[pipeline_id] = {}

            for i in range(parser_count):
                if pipeline_dict['value'][i]['createdDate'][0:10] not in self.pipeline_date_list:
                    logger.debug("pipeline_id {} parser_count {}; current {} not in date list {} ".format(pipeline_id, parser_count, pipeline_dict['value'][i]['createdDate'][0:10], self.pipeline_date_list))
                    break

                build_id = pipeline_dict['value'][i]['id']
                build_url = "https://dev.azure.com/mssonic/internal/_build/results?buildId=" + str(pipeline_dict['value'][i]['id'])

                # logger.info("build index {} build_id {}".format(i, build_id))

                if pipeline_dict['value'][i]["state"] == "inProgress":
                    self.pipeline_build_result_dict[pipeline_id][build_id] = {  'name' : pipeline_dict['value'][i]['name'], 
                                                                                'id' : pipeline_dict['value'][i]['id'], 
                                                                                'state' : pipeline_dict['value'][i]['state'], 
                                                                                'result' : None, 
                                                                                'url' : build_url, 
                                                                                'createdDate' : pipeline_dict['value'][i]['createdDate'][0:19], 
                                                                                'finishedDate' : 'inProgress' ,
                                                                                'build_details' : {'Lock Testbed' : 'NA', 
                                                                                                    'Upgrade Image' : 'NA', 
                                                                                                    'Deploy Minigraph' : 'NA', 
                                                                                                    'Run Tests' : 'NA', 
                                                                                                    'Other failure' : 'NA'}}
                elif pipeline_dict['value'][i].get('result'):
                    # logger.info("i {} date : {} {}".format(i, pipeline_dict['value'][i]['createdDate'], type(pipeline_dict['value'][i]['createdDate'])))
                    # logger.info("i {} value {}".format(i, pipeline_dict['value'][i]))

                    start_time = datetime.datetime.strptime(pipeline_dict['value'][i]['createdDate'][0:19], '%Y-%m-%dT%H:%M:%S')
                    finish_time = datetime.datetime.strptime(pipeline_dict['value'][i]['finishedDate'][0:19], '%Y-%m-%dT%H:%M:%S')
                    run_time = finish_time - start_time

                    self.pipeline_build_result_dict[pipeline_id][build_id] = {  'name' : pipeline_dict['value'][i]['name'], 
                                                                                'id' : pipeline_dict['value'][i]['id'], 
                                                                                'state' : pipeline_dict['value'][i]['state'], 
                                                                                'result' : pipeline_dict['value'][i]['result'], 
                                                                                'url' : build_url, 
                                                                                'createdDate' : pipeline_dict['value'][i]['createdDate'][0:19], 
                                                                                'finishedDate' : "{} \n run time: {}".format(pipeline_dict['value'][i]['finishedDate'][0:19], run_time),
                                                                                'build_details' : {'Lock Testbed' : 'NA', 
                                                                                                    'Upgrade Image' : 'NA', 
                                                                                                    'Deploy Minigraph' : 'NA', 
                                                                                                    'Run Tests' : 'NA', 
                                                                                                    'Other failure' : 'NA'}}

            logger.info("pipeline id {} pipeline_build_result_dict {}".format(pipeline_id, self.pipeline_build_result_dict[pipeline_id]))
        else:
            logger.error("trigger pipeline build failed: code {} ".format(build_response.status_code))

        return




    def parser_pipeline_status_result_by_branch(self, branch):
        if 'master' == branch :
            if self.nightly_pipelines_dict.get('master', None):
                for index, index_info in self.nightly_pipelines_dict['master'].items():
                    self.parser_pipeline_status_result_by_pipeline_id(index_info['pipeline_id'])
                # logger.info("master pipeline_build_result_dict {}".format(self.pipeline_build_result_dict))
        elif 'internal' == branch :
            if self.nightly_pipelines_dict.get('internal', None):
                for index, index_info in self.nightly_pipelines_dict['internal'].items():
                    self.parser_pipeline_status_result_by_pipeline_id(index_info['pipeline_id'])
                # logger.info("internal pipeline_build_result_dict {}".format(self.pipeline_build_result_dict))
        elif 'internal-202012' == branch :
            if self.nightly_pipelines_dict.get('internal-202012', None):
                for index, index_info in self.nightly_pipelines_dict['internal-202012'].items():
                    self.parser_pipeline_status_result_by_pipeline_id(index_info['pipeline_id'])
                # logger.info("internal-202012 pipeline_build_result_dict {}".format(self.pipeline_build_result_dict))
        elif 'internal-202205' == branch :
            if self.nightly_pipelines_dict.get('internal-202205', None):
                for index, index_info in self.nightly_pipelines_dict['internal-202205'].items():
                    self.parser_pipeline_status_result_by_pipeline_id(index_info['pipeline_id'])
                # logger.info("internal-202205 pipeline_build_result_dict {}".format(self.pipeline_build_result_dict))
        else:
            logger.error("ERROR: branch {} mismatch ".format(branch))


    def parser_pipeline_status_result(self):
        TOKEN = os.environ.get('AZURE_DEVOPS_MSSONIC_TOKEN')
        if not TOKEN:
            logger.error("Get token failed, Must export environment variable AZURE_DEVOPS_MSSONIC_TOKEN")
            if self.verbose :
                logger.error("No token, move all testbeds {} to custo table".format("None"))
        AUTH = ('', TOKEN)

        # collect nightly pipeline build info first, need using pipeline ID to collect build result
        self.collect_nightly_pipeline_build()

        # logger.info("nightly_pipelines_dict {}  ".format(self.nightly_pipelines_dict))
        self.parser_pipeline_status_result_by_branch('master')
        self.parser_pipeline_status_result_by_branch('internal')
        self.parser_pipeline_status_result_by_branch('internal-202012')
        self.parser_pipeline_status_result_by_branch('internal-202205')

        logger.info("pipeline_build_result_dict {}".format(self.pipeline_build_result_dict))


    def collect_pipeline_build_test_result_by_build_id(self, build_id):
        """
        query kusto to get the test result based on build ID
        """

        curr_build_test_result = self.kusto_checker.query_build_test_result(build_id)

        if self.build_test_case_result_dict.get(build_id, None):
            logger.info("build_id {} has already had result".format(build_id))
            del self.build_test_case_result_dict[build_id]

        self.build_test_case_result_dict[build_id] = {}
        self.build_test_case_result_dict[build_id]['success'] = []
        self.build_test_case_result_dict[build_id]['failure'] = []
        self.build_test_case_result_dict[build_id]['error'] = []
        for row in curr_build_test_result.primary_results[0].rows:
            # logger.info("TestbedName {} FullTestPath {} Result {} Runtime {} Summary {}".format(row['TestbedName'], row['FullTestPath'], row['Result'], row['Runtime'], row['Summary']))
            self.build_test_case_result_dict[build_id][row['Result']].append(row['FullTestPath'])

        logger.info("build_test_case_result_dict {} ".format(self.build_test_case_result_dict[build_id]))
        logger.info("build id {} success {} failure {} err {} ".format(build_id, len(self.build_test_case_result_dict[build_id]['success']), 
                                                                    len(self.build_test_case_result_dict[build_id]['failure']), len(self.build_test_case_result_dict[build_id]['error'])))

    def create_branch_pipeline_list(self, branch_name):
        nightly_pipeline_dict = {}
        count = 1

        for _, pipeline in self.pipeline_parser_analyzer_dict.items():
            for branch, branch_pipeline in pipeline.items():
                if branch == branch_name:
                    for _, branch_pipeline_info in branch_pipeline.items():
                        logger.debug("branch {} : {}".format(branch, branch_pipeline_info)) 
                        if '-t2-' not in branch_pipeline_info['pipeline_name']:
                            nightly_pipeline_dict[count] = branch_pipeline_info['pipeline_name']
                            count += 1

        logger.info("branch_name {} nightly_pipeline_dict {}".format(branch_name, nightly_pipeline_dict)) 
        return nightly_pipeline_dict


    def excel_create_file(self):
        logger.info("save_file_name {}".format(self.save_file_name)) 
        if os.path.exists(self.save_file_name):
            os.system("rm -f {}".format(self.save_file_name))

        wb = Workbook()
        
        self.curr_row = 1
        self.curr_col = 1
        self.excel_sheet_create(wb, 'master', nightly_pipeline_master)
        self.curr_row = 1
        self.curr_col = 1
        self.excel_sheet_create(wb, 'internal', nightly_pipeline_internal)
        self.curr_row = 1
        self.curr_col = 1
        self.excel_sheet_create(wb, 'internal-202205', nightly_pipeline_202205)
        self.curr_row = 1
        self.curr_col = 1
        self.excel_sheet_create(wb, 'internal-202012', nightly_pipeline_202012)

        wb.save(self.save_file_name)

        return


    def build_id_result_excel_sheet_create(self, wb, build_id, build_result_dict):
        logger.info("excel_sheet_create {} - {}".format(self.save_file_name, build_id)) 

        self.curr_row = 1
        self.curr_col = 1

        ws_sheet = wb.create_sheet(str(build_id))

        # add date
        ws_sheet.cell(row = self.curr_row, column = self.curr_col,     value = "Date: " + str(datetime.datetime.now()))
        # ws_sheet.cell(row = self.curr_row, column = self.curr_col + 1, value = str(datetime.datetime.now()))
        self.curr_row += 2

        save_row = self.curr_row
        
        for value in self.result_value_list:
            logger.info("{} cases count: {} ".format(value, len(build_result_dict[value].keys())))
            ws_sheet.cell(row = self.curr_row, column = self.curr_col, value = "{} ".format(value))      
            ws_sheet.cell(row = self.curr_row, column = self.curr_col + 1, value = "{} ".format(len(build_result_dict[value].keys())))
            ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].alignment = Alignment(horizontal='right', vertical='bottom', wrap_text=True)
            self.curr_row += 1
            self.curr_col = 1

        success_case_count = len(build_result_dict['success'].keys())
        total_cases_count = success_case_count + len(build_result_dict['error'].keys()) + len(build_result_dict['failure'].keys())
        if total_cases_count == 0:
            success_rate = 0
        else:
            logger.info("success rate {:.2%} ".format(success_case_count / total_cases_count))
            success_rate = success_case_count / total_cases_count
        ws_sheet.cell(row = save_row, column = self.curr_col + 2, value = "{} ".format('success rate'))      
        ws_sheet.cell(row = save_row, column = self.curr_col + 2 + 1, value = "{:.2%} ".format(success_rate))
        ws_sheet[get_column_letter(save_row)+str(self.curr_col + 2 + 1)].alignment = Alignment(horizontal='right', vertical='bottom', wrap_text=True)
        
        self.curr_row += 1
        self.curr_col = 1

        logger.info("##### build_result_dict {}".format(build_result_dict.items()))

        # keys = ["StartTimeUTC", "TestbedName", "OSVersion", "StartTime", "Runtime", "Result", "BuildId", "ModulePath", "TestCase", "FullTestPath", "Summary"]
        for result_status, result_info in build_result_dict.items():
            if result_status == 'success' or result_status == 'skipped':
                continue
            for index, case in result_info.items():
                for i, key in enumerate(self.keys):
                    value = case.get(key, "")
                    # logger.debug("i {} item_key {}, value {}".format(i, key, value)) 
                    ws_sheet.cell(row = self.curr_row, column = self.curr_col + i, value = value)
                self.curr_row += 1
                self.curr_col = 1

        return
    

    def create_pipeline_build_case_result_xls_file(self):
        logger.info("save_file_name {}".format(self.save_file_name)) 

        name, ext = self.save_file_name.rsplit('.', 1)
        new_name = '{}-{}'.format(name, 'build_result')
        pipeline_case_result_file = '{}.{}'.format(new_name, ext)

        
        if os.path.exists(pipeline_case_result_file):
            os.system("rm -f {}".format(pipeline_case_result_file))

        wb = Workbook()

        for build_id in sorted(self.pipeline_build_case_result.keys()):
            logger.info("build_id {} ".format(build_id)) 
            self.build_id_result_excel_sheet_create(wb, build_id, self.pipeline_build_case_result[build_id] )

        wb.save(pipeline_case_result_file)
        return


    def excel_sheet_create(self, wb, branch, nightly_pipeline_dict):
        logger.info("excel_sheet_create {} - {}".format(self.save_file_name, branch)) 

        pipeline_save_row = 0
        pipeline_save_col = 0
        pipeline_date_save_row = 0
        pipeline_date_save_col = 0
        pipeline_date_build_save_row = 0
        pipeline_date_build_save_col = 0

        ws_sheet = wb.create_sheet(branch) 

        # add date
        ws_sheet.cell(row = self.curr_row, column = self.curr_col,     value = "Date ")
        ws_sheet.cell(row = self.curr_row, column = self.curr_col + 1, value = str(datetime.datetime.now()))
        self.curr_row += 2

        # add headline
        self.curr_col = 3
        for i in range(len(self.pipeline_date_list)):
            ws_sheet.cell(row = self.curr_row, column = self.curr_col, value = self.pipeline_date_list[i])
            self.curr_col += self.pipeline_cols

        self.curr_row += 1
        self.curr_col = 1

        # # self.curr_row = row_pipeline_start_save = self.curr_row
        pipeline_save_row = self.curr_row
        pipeline_save_col = self.curr_col
        pipeline_save_row_max = pipeline_save_row

        nightly_pipeline_dict = self.sorted_key_dict(nightly_pipeline_dict)
        logger.info("branch {} nightly_pipeline_dict {} ".format(branch, nightly_pipeline_dict))

        first_pipeline = True

        # import pdb; pdb.set_trace()
        for pipeline_index in nightly_pipeline_dict.keys():
            logger.debug("branch {} pipeline_index {} ".format(branch, pipeline_index))
            if not self.nightly_pipelines_dict[branch].get(pipeline_index):
                logger.debug("branch {} pipeline_index {} has no such pipeline ".format(branch, pipeline_index))
                continue
            pipeline_id = (self.nightly_pipelines_dict[branch][pipeline_index]["pipeline_id"])

            logger.debug("branch {} pipeline_index {} pipeline_id {} ".format(branch, pipeline_index, pipeline_id))
            if pipeline_id == None:
                logger.debug("branch {} pipeline_index {} pipeline_id None, continue ".format(branch, pipeline_index))
                continue
            else:
                pipeline_results = self.pipeline_build_result_dict[pipeline_id]
                if pipeline_results == None:
                    logger.debug("branch {} pipeline_index {} pipeline_id {} pipeline_results None, continue ".format(branch, pipeline_index, pipeline_id))
                    continue

            self.curr_col = 1

            logger.debug("### pipeline {},  [{}][{}] ".format(pipeline_index, self.curr_row, self.curr_col))

            if first_pipeline:
                first_pipeline = False
            else:
                self.curr_row = pipeline_save_row_max + self.pipeline_rows

            # self.pipeline_rows = 14
            pipeline_save_row = self.curr_row
            pipeline_save_col = self.curr_col

            testbed_name = self.nightly_pipelines_dict[branch][pipeline_index]['testbed_name']
            schedules = {}
            for key in ["master", "internal", "internal-202205", "internal-202012", "internal-201911"]:
                data = self.pipeline_parser_analyzer_dict[testbed_name].get(key, {})
                schedules_tmp = "\n".join(yml_info.get("schedule", "") for yml_info in data.values())
                schedules[key] = schedules_tmp
                logger.debug("data {} schedules[key]  {} ".format(data, schedules[key]))

            pipeline_detail_head = { 
                                        "index" : pipeline_index, 
                                        "pipeline_name" : self.nightly_pipelines_dict[branch][pipeline_index]['pipeline_name'],
                                        "pipeline_id" : self.nightly_pipelines_dict[branch][pipeline_index]['pipeline_id'],
                                        "testbed_name" : self.nightly_pipelines_dict[branch][pipeline_index]['testbed_name'],
                                        "schedule" : self.nightly_pipelines_dict[branch][pipeline_index]['schedule'],
                                        "image_url" : self.nightly_pipelines_dict[branch][pipeline_index]['image_url'],

                                        "schedule-int" : schedules.get("internal", ""),
                                        "schedule-master" : schedules.get("master", ""),
                                        "schedule-202205" : schedules.get("internal-202205", ""),
                                        "schedule-202012" : schedules.get("internal-202012", ""),
                                        "schedule-201911" : schedules.get("internal-221911", ""),
                                    }
            logger.debug("pipeline_index {} {}".format(pipeline_index, pipeline_detail_head)) 

            for item_name, item_value in pipeline_detail_head.items() :
                ws_sheet.cell(row = self.curr_row, column = self.curr_col, value = item_name)
                ws_sheet.cell(row = self.curr_row, column = self.curr_col + 1, value = item_value)

                ws_sheet.column_dimensions[get_column_letter(self.curr_col)].width = 15
                ws_sheet.column_dimensions[get_column_letter(self.curr_col + 1)].width = 45
                ws_sheet[get_column_letter(self.curr_col)+str(self.curr_row)].alignment = Alignment(horizontal='left', vertical='bottom')
                ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].alignment = Alignment(horizontal='left', vertical='bottom')


                ws_sheet[get_column_letter(self.curr_col)+str(self.curr_row)].alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
                ws_sheet[get_column_letter(self.curr_col)+str(self.curr_row)].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))                            
                ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
                ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))


                if item_name == 'index':
                    ws_sheet[get_column_letter(self.curr_col)+str(self.curr_row)].fill = PatternFill("solid", fgColor='FFA500')
                    ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].fill = PatternFill("solid", fgColor='FFA500')
                elif "schedule-" in item_name:
                    ws_sheet[get_column_letter(self.curr_col)+str(self.curr_row)].fill = PatternFill("solid", fgColor='C0C0C0')
                    ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].fill = PatternFill("solid", fgColor='C0C0C0')

                self.curr_row += 1
                self.curr_col = 1

     
            # add build result
            self.curr_row = pipeline_save_row
            self.curr_col = 3

            pipeline_date_save_row = self.curr_row
            pipeline_date_save_col = self.curr_col

            logger.debug("### pipeline_results  {} ".format(pipeline_results))
            logger.debug("### build index  [{}][{}] ".format(self.curr_row, self.curr_col))


            current_date = self.pipeline_date_list[0]
            count_in_current_date = 0
            logger.debug("### current_date  {} ".format(current_date))
            first_item = True

            logger.debug("### build index  [{}][{}], {} ".format(self.curr_row, self.curr_col, pipeline_save_row_max))
            # if build result is empty
            if not bool(pipeline_results):
                pipeline_save_row_max = self.curr_row
                # logger.info("### build index  [{}][{}] ".format(self.curr_row, self.curr_col))
                continue

            for build_id, build_info in pipeline_results.items():
                logger.debug("branch {} pipeline_index {} pipeline_id {} build_id {} ".format(branch, pipeline_index, pipeline_id, build_id))
                logger.debug("### pipeline {} build index {},  [{}][{}] ".format(pipeline_index, build_id, self.curr_row, self.curr_col))

                if build_info['createdDate'][0:10] != current_date:
                    logger.debug("branch {} pipeline_index {} pipeline_id {} build_id {} build_date {} current_date {}".format(branch, pipeline_index, pipeline_id, build_id, build_info['createdDate'][0:10], current_date))
                    current_date = build_info['createdDate'][0:10] 
                    count_in_current_date = 0
                else:
                    if current_date == self.pipeline_date_list[0] and first_item == True:
                        logger.debug("branch {} pipeline_index {} pipeline_id {} build_id {} build_date {} current_date {}".format(branch, pipeline_index, pipeline_id, build_id, build_info['createdDate'][0:10], current_date))
                        count_in_current_date = 0
                        first_item = False
                    else:
                        logger.debug("branch {} pipeline_index {} pipeline_id {} build_id {} build_date {} current_date {}".format(branch, pipeline_index, pipeline_id, build_id, build_info['createdDate'][0:10], current_date))
                        count_in_current_date += 1

                diff_days = (datetime.datetime.today() - datetime.datetime.strptime(build_info['createdDate'][0:10], "%Y-%m-%d")).days
                logger.debug("### build day {} diff {}".format(build_info['createdDate'][0:10], diff_days))

                self.curr_row = pipeline_save_row + count_in_current_date * self.pipeline_rows
                if pipeline_save_row_max < self.curr_row:
                    logger.debug("branch {} pipeline_index {} pipeline_id {} build_id {}; max {} -> {} ".format(branch, pipeline_index, pipeline_id, build_id, pipeline_save_row_max, self.curr_row))
                    pipeline_save_row_max = self.curr_row
                self.curr_col = 3 + diff_days * self.pipeline_cols

                logger.debug("### pipeline {} build index {} current_date {}, [{}][{}] ".format(pipeline_index, build_id, current_date, self.curr_row, self.curr_col))

                build_result_list = ['name', 'id', 'state', 'result', 'url', 'createdDate', 'finishedDate']
                for i in range(len(build_result_list)):
                    # logger.info("add [{}][{}] : {} - {}".format(self.curr_row, self.curr_col, build_result_list[i], result[build_result_list[i]]))
                    ws_sheet.cell(row = self.curr_row, column = self.curr_col,     value = build_result_list[i])
                    ws_sheet.cell(row = self.curr_row, column = self.curr_col + 1, value = build_info[build_result_list[i]])

                    ws_sheet.column_dimensions[get_column_letter(self.curr_col)].width = self.pipeline_cols_name_width
                    ws_sheet.column_dimensions[get_column_letter(self.curr_col + 1)].width = self.pipeline_cols_data_width
                    ws_sheet[get_column_letter(self.curr_col)+str(self.curr_row)].alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
                    ws_sheet[get_column_letter(self.curr_col)+str(self.curr_row)].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))
                    ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
                    ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

                    if build_result_list[i] == 'state':
                        if 'inProgress' in build_info[build_result_list[i]] :
                            # ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].fill = PatternFill("solid", fgColor='00ff00')
                            ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].font = Font(size=12, color='FF0000')
                    elif build_result_list[i] == 'result' and build_info.get('result'):
                        if 'failed' in build_info[build_result_list[i]] :
                            # ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].fill = PatternFill("solid", fgColor='FF0000')
                            ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].font = Font(size=12, color='FF0000')

                    self.curr_row += 1
                    # self.curr_col = col_save

                build_result_details_list = ['Lock Testbed', 'Upgrade Image', 'Deploy Minigraph', 'Run Tests', 'Other failure']
                for i in range(len(build_result_details_list)):
                    # logger.info("add [{}][{}] : {} - {}".format(self.curr_row, self.curr_col, build_result_details_list[i], result['build_details'][build_result_details_list[i]]))
                    ws_sheet.cell(row = self.curr_row, column = self.curr_col,     value = build_result_details_list[i])
                    ws_sheet.cell(row = self.curr_row, column = self.curr_col + 1, value = build_info['build_details'][build_result_details_list[i]])

                    ws_sheet.column_dimensions[get_column_letter(self.curr_col)].width = self.pipeline_cols_name_width
                    ws_sheet.column_dimensions[get_column_letter(self.curr_col + 1)].width = self.pipeline_cols_data_width
                    ws_sheet[get_column_letter(self.curr_col)+str(self.curr_row)].alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
                    ws_sheet[get_column_letter(self.curr_col)+str(self.curr_row)].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))                            
                    ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
                    ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

                    # logger.debug("### build index  [{}][{}] build_result_list {} {} ".format(self.curr_row, self.curr_col, build_result_details_list[i], result['build_details'][build_result_details_list[i]]))

                    # if result['build_details'][build_result_details_list[i]] == 'failed':
                    if 'failed' in build_info['build_details'][build_result_details_list[i]] or 'canceled' in build_info['build_details'][build_result_details_list[i]]:
                        # ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].fill = PatternFill("solid", fgColor='FF0000')
                        ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].font = Font(size=12, color='FF0000')
                    elif 'pass rate' in build_info['build_details'][build_result_details_list[i]] :
                        # ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].fill = PatternFill("solid", fgColor='00ff00')
                        ws_sheet[get_column_letter(self.curr_col + 1)+str(self.curr_row)].font = Font(size=12, color='00B050')

                    self.curr_row += 1
                    # self.curr_col = col_save

        return


    def create_testbed_cfg(self):
        logger.info("create_testbed_cfg {}".format(self.save_file_name)) 
        pipeline_testbeds = pipeline_analyzer.pipeline_parser_analyzer_dict.keys()
        logger.info("pipeline_testbeds {}".format(pipeline_testbeds))

        device_group = []

        with open(os.path.join(ANSIBLE_DIR, 'testbed.yaml')) as f:
            logger.info("parser testbed.yaml")
            testbed_dict = yaml.safe_load(f)
        logger.info("testbed_dict {} {}".format(type(testbed_dict), len(testbed_dict)))
        logger.info("testbed_dict {}".format(testbed_dict))

        # file_list = ['str', 'str3', 'str3', 'strsvc', 'bjw']
        file_list = ['strsvc']
        for file in file_list:
            with open(os.path.join(ANSIBLE_DIR, file)) as f:
                logger.info("parser {}".format(file))
                parser_dict = yaml.safe_load(f)
                logger.info("parser_dict {}".format(parser_dict))
                if parser_dict.get('sonic') and parser_dict['sonic'].get('children') :
                    device_group = list(set(device_group + list(parser_dict['sonic']['children'].keys())))
                    logger.info("device_group {}".format(device_group))

        for group in device_group:
            logger.info("group {}".format(group))
            for file in file_list:
                with open(os.path.join(ANSIBLE_DIR, file)) as f:
                    logger.info("parser {}".format(file))
                    parser_dict = yaml.safe_load(f)
                    logger.info("parser_dict {}".format(parser_dict))
                    if parser_dict.get('sonic') and parser_dict['sonic'].get('children') :
                        device_group = list(set(device_group + list(parser_dict['sonic']['children'].keys())))
                        logger.info("device_group {}".format(device_group))

        return


if __name__ == '__main__':

    parser = argparse.ArgumentParser(description='Completeness level')
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        description="nightly hawk pipeline analyzer")

    parser.add_argument(
        '-v', '--verbose', help='Set verbose output level', type=str,
        required=False, default='INFO'
    )

    parser.add_argument(
        '-d', '--days2parser', help='input the number of days need to parser', type=int,
        choices=range(1,60), required=False, default=3
    )

    parser.add_argument(
        '-c', '--useCache', help='using cache, skip collecting pipeline information', action='store_true',
        required=False
    )

    parser.add_argument(
        '-p', '--pipelineCustomer', help='use customer pipeline config to parser', action='store_false',
        required=False, default=True
    )

    args = parser.parse_args()
    logger.info("Parser {} days pipeline result, using logging lever {}, using cache {}, pipeline customer {}".format(args.days2parser, args.verbose, args.useCache, args.pipelineCustomer))


    pipeline_analyzer = Nightly_hawk_pipeline_analyzer(verbose = args.verbose, days2parser = args.days2parser)

    skip_parser_pipeline_info = args.useCache
    if skip_parser_pipeline_info and os.path.exists('pipeline_parser_dict_debug.json'):
        with open('pipeline_parser_dict_debug.json') as f:
            logger.info("parser_pipeline_info using cache file")
            pipeline_analyzer.pipeline_parser_analyzer_dict = json.load(f)
    else:
        logger.info("parser_pipeline_info online ")
        pipeline_analyzer.pipeline_parser_analyzer_dict = pipeline_analyzer.nightly_pipeline_check.collect_nightly_build_pipelines()

    # logger.info("pipeline_parser_analyzer_dict {} ".format(pipeline_analyzer.pipeline_parser_analyzer_dict))

    if not args.pipelineCustomer:
        nightly_pipeline_internal.clear()
        nightly_pipeline_master.clear()
        nightly_pipeline_202012.clear()
        nightly_pipeline_202205.clear()
        nightly_pipeline_internal = pipeline_analyzer.create_branch_pipeline_list('internal')
        nightly_pipeline_master = pipeline_analyzer.create_branch_pipeline_list('master')
        nightly_pipeline_202205 = pipeline_analyzer.create_branch_pipeline_list('internal-202205')
        nightly_pipeline_202012 = pipeline_analyzer.create_branch_pipeline_list('internal-202012')
        logger.info("nightly_pipeline_internal {} ".format(nightly_pipeline_internal))
        logger.info("nightly_pipeline_master {} ".format(nightly_pipeline_master))
        logger.info("nightly_pipeline_202205 {} ".format(nightly_pipeline_202205))
        logger.info("nightly_pipeline_202012 {} ".format(nightly_pipeline_202012))

    logger.info("parser_pipeline_status_result ")
    pipeline_analyzer.parser_pipeline_status_result()

    logger.info("parser_pipeline_build_result ")
    pipeline_analyzer.parser_pipeline_build_result()
    
    logger.info("excel_create_file ")
    pipeline_analyzer.excel_create_file()

    logger.info("create case result file ")
    pipeline_analyzer.create_pipeline_build_case_result_xls_file()

    logger.info("Nightly_hawk_pipeline_analyzer complete ")

    


